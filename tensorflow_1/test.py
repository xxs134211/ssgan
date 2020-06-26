import numpy as np
from openpyxl import load_workbook


def write_excel(path, learn_rate, model_number, accuracy):
    wb = load_workbook(path)
    if model_number == 1:
        ws = wb["model_1"]
    else:
        ws = wb["model_2"]
    ws['D1'] = 'lr=' + format(learn_rate)
    for i in range(len(accuracy)):
        ws.cell(row=i+2, column=4).value = accuracy[i]
    # for hang in range(len(accuracy)):
    #     for i in accuracy:
    #         h = hang + 1
    #         ws.cell(row=h, column=2).value = accuracy[hang].value
    wb.save(path)


Path = 'D:/python/ssgan_tf2.0/accuracy.xlsx'
Learn_rate = 0.001
Model_number = 2
acc = [1, 3, 2, 1, 4, 5, 6, 7, 8.1, 23, 243252, 234]
write_excel(Path, Learn_rate, Model_number, acc)
print(acc[1])
