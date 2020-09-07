import os
import datetime

import numpy as np
import tensorflow as tf
from openpyxl import load_workbook
from tensorflow import keras
import ssgan_dataset_tf2
from model_2.ssgan_model_2_tf2 import Generator, Discriminator
import matplotlib.pyplot as plt
import time
import warnings

warnings.filterwarnings("ignore")
# os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
devices = tf.config.experimental.list_physical_devices('GPU')
tf.config.experimental.set_memory_growth(devices[0], True)


# 准备与真实数据的标签相乘的二进制标签掩码
def get_labeled_mask(labeled_rate, batch_size):
    labeled_mask = np.zeros([batch_size], dtype=np.float32)
    labeled_count = np.int(batch_size * labeled_rate)
    labeled_mask[range(labeled_count)] = 1.0
    np.random.shuffle(labeled_mask)
    return labeled_mask


# ########### 扩展标签，加入第n+1类标签，三类轴承数据 ############
def prepare_extended_label(label):
    # add extra label for fake data
    extended_label = tf.concat([tf.zeros([tf.shape(label)[0], 1]), label], axis=1)

    return extended_label


def d_loss_fn(generator, discriminator, batch_z, batch_x, labeled_mask, extended_label, is_training):
    fake_images = generator(batch_z, is_training)
    D_fake_features, D_fake_logits, D_fake_prob, D_fake_mid = discriminator(fake_images, is_training)
    D_real_features, D_real_logits, D_real_prob, D_real_mid = discriminator(batch_x, is_training)

    temp = tf.nn.softmax_cross_entropy_with_logits(logits=D_real_logits, labels=extended_label)
    D_L_Supervised = tf.reduce_sum(tf.multiply(temp, labeled_mask)) / tf.reduce_sum(labeled_mask)
    D_L_RealUnsupervised = tf.reduce_mean(tf.nn.sigmoid_cross_entropy_with_logits(
        logits=D_real_logits[:, 0], labels=tf.zeros_like(D_real_logits[:, 0], dtype=tf.float32)))
    D_L_FakeUnsupervised = tf.reduce_mean(tf.nn.sigmoid_cross_entropy_with_logits(
        logits=D_fake_logits[:, 0], labels=tf.ones_like(D_fake_logits[:, 0], dtype=tf.float32)))
    data_moments = tf.reduce_mean(D_real_mid, axis=0)
    sample_moments = tf.reduce_mean(D_fake_mid, axis=0)
    D_L_2 = tf.reduce_mean(tf.square(data_moments - sample_moments))

    D_L = D_L_Supervised + D_L_RealUnsupervised + D_L_FakeUnsupervised
    return D_L


def g_loss_fn(generator, discriminator, batch_z, batch_x, is_training):
    # G_L_1 -> Fake data wanna be real
    fake_images = generator(batch_z, is_training)
    D_fake_features, D_fake_logits, D_fake_prob, D_fake_mid = discriminator(fake_images, is_training)
    D_real_features, D_real_logits, D_real_prob, D_real_mid = discriminator(batch_x, is_training)
    G_L_1 = tf.reduce_mean(tf.nn.sigmoid_cross_entropy_with_logits(
        logits=D_fake_logits[:, 0], labels=tf.zeros_like(D_fake_logits[:, 0], dtype=tf.float32)))

    # G_L_2 -> Feature matching
    data_moments = tf.reduce_mean(D_real_features, axis=0)
    sample_moments = tf.reduce_mean(D_fake_features, axis=0)
    G_L_2 = tf.reduce_mean(tf.square(data_moments - sample_moments))

    # G_L_3 -> Feature matching
    data_moments_mid = tf.reduce_mean(D_real_mid, axis=0)
    sample_moments_mid = tf.reduce_mean(D_fake_mid, axis=0)
    G_L_3 = tf.reduce_mean(tf.square(data_moments_mid - sample_moments_mid))

    G_L = G_L_1 + G_L_2 + 0.5 * G_L_3

    return G_L


def accuracy(discriminator, batch_x, extended_label, is_training):
    D_real_features, D_real_logits, D_real_prob, D_real_mid = discriminator(batch_x, is_training)
    prediction_value = tf.argmax(D_real_prob[:, 1:], 1)
    prediction = tf.equal(tf.argmax(D_real_prob[:, 1:], 1),
                          tf.argmax(extended_label[:, 1:], 1))
    acc = tf.reduce_mean(tf.cast(prediction, tf.float32))
    return acc, prediction_value


def Draw(hist, name, epoch, show=False, save=False, is_loss=True):
    plt.figure()
    Time = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
    if is_loss:
        plt.plot(hist['G_losses'], 'b', label='generator')
        plt.plot(hist['D_losses'], 'r', label='discriminator')
        plt.xlabel('Epoch')
        plt.ylabel('Loss')
        plt.legend()
        if save:
            if not os.path.exists('Loss'):
                os.mkdir('Loss')
            plt.savefig("Loss/GAN_loss_lr[{}]epoch[{}]time[{}].png".format(name, epoch, Time))
    else:
        plt.plot(hist, 'b', label='acc')
        plt.xlabel('Epoch')
        plt.ylabel('Accuracy')
        plt.legend()
        if save:
            if not os.path.exists('plot'):
                os.mkdir('plot')
            plt.savefig("plot/GAN_acc_lr[{}]epoch[{}]time[{}].png".format(name, epoch, Time))
    if show:
        plt.show()
    else:
        plt.close()


# 操作Excel表格
def write_excel(path, learn_rate, model_number, Accuracy):
    wb = load_workbook(path)
    Time = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
    if model_number == 1:
        ws = wb["model_1"]
    else:
        ws = wb["model_2"]
    ws['V1'] = 'lr=' + format(learn_rate) + Time
    for i in range(len(Accuracy)):
        ws.cell(row=i + 2, column=22).value = Accuracy[i]
    wb.save(path)


def main(learning_rate, epochs):
    batch_size = 64
    # learning_rate = 0.0002
    z_dim = 100
    is_training = True
    # epochs = 2
    labeled_rate = 0.2
    Train_acc = []
    test_acc = []
    train_hist = {'D_losses': [], 'G_losses': []}
    file = ['', '', '']  # 模型文件名称，删除之前保存的文件名称

    generator = Generator()
    generator.build(input_shape=(1, z_dim))
    discriminator = Discriminator()
    discriminator.build(input_shape=(None, 32, 32, 1))

    g_optimizer = keras.optimizers.Adam(learning_rate=learning_rate, beta_1=0.5)
    d_optimizer = keras.optimizers.Adam(learning_rate=learning_rate, beta_1=0.5)
    discriminator.summary()
    Time = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    path = 'D:/python/ssgan_tf2.0/log_dir'
    log_dir = os.path.join(path, Time)
    if not os.path.exists(log_dir):
        os.mkdir(log_dir)
    writer = tf.summary.create_file_writer(log_dir)

    no_of_batches = int(ssgan_dataset_tf2.train_X.shape[0] / batch_size) + 1
    for epoch in range(epochs):
        train_accuracies, train_D_losses, train_G_losses = [], [], []
        for i in range(no_of_batches - 1):
            # 准备训练数据
            batch_x = ssgan_dataset_tf2.train_X[i * batch_size:batch_size + i * batch_size, ]
            batch_label = ssgan_dataset_tf2.train_Y[i * batch_size:batch_size + i * batch_size, ]
            batch_reshaped = batch_x.reshape([-1, 32, 32, 1])
            batch_z = np.random.normal(0, 1, (batch_size, 1, 1, z_dim))
            mask = get_labeled_mask(labeled_rate, batch_size)
            extended_label = prepare_extended_label(batch_label)

            # 准备验证数据
            valid_data = ssgan_dataset_tf2.valid_X
            valid_label = ssgan_dataset_tf2.valid_Y
            valid_data_reshaped = valid_data.reshape([-1, 32, 32, 1])
            valid_extended_label = prepare_extended_label(valid_label)

            # 判别器前向计算
            with tf.GradientTape() as tape:
                d_loss = d_loss_fn(generator, discriminator, batch_z, batch_reshaped, mask, extended_label, is_training)
                grads = tape.gradient(d_loss, discriminator.trainable_variables)
                d_optimizer.apply_gradients(zip(grads, discriminator.trainable_variables))

            with tf.GradientTape() as tape:
                g_loss = g_loss_fn(generator, discriminator, batch_z, batch_reshaped, is_training)
                grads = tape.gradient(g_loss, generator.trainable_variables)
                g_optimizer.apply_gradients(zip(grads, generator.trainable_variables))

            train_accuracy, _ = accuracy(discriminator, valid_data_reshaped, valid_extended_label, None)
            train_accuracies.append(train_accuracy)
            print('Epoch [{}]/[{}]'.format(epoch, epochs), 'Batch evaluated [{}]/[{}]'.format(str(i + 1),
                                                                                              no_of_batches - 1))

            train_D_losses.append(d_loss)
            train_G_losses.append(g_loss)
        tr_GL = np.mean(train_G_losses)
        tr_DL = np.mean(train_D_losses)
        tr_acc = np.mean(train_accuracies)
        train_hist['D_losses'].append(tr_DL)
        train_hist['G_losses'].append(tr_GL)
        Train_acc.append(tr_acc)
        print('After epoch: ' + str(epoch + 1) + ' Generator loss: '
              + str(tr_GL) + ' Discriminator loss: ' + str(tr_DL) + ' Accuracy: ' + str(tr_acc))

        with writer.as_default():
            tf.summary.scalar("train/tr_DL", tr_DL, epoch)
            tf.summary.scalar("train/tr_GL", tr_GL, epoch)
            tf.summary.scalar("train/tr_acc", tr_acc, epoch)

        # 准备测试数据
        test_data = ssgan_dataset_tf2.test_X
        test_label = ssgan_dataset_tf2.test_Y
        test_data_reshaped = test_data.reshape([-1, 32, 32, 1])
        test_extended_label = prepare_extended_label(test_label)
        test_accuracy, _ = accuracy(discriminator, test_data_reshaped, test_extended_label, False)
        print('测试集：' + str(test_accuracy.numpy()))
        epoch_accuracy = test_accuracy
        with writer.as_default():
            tf.summary.scalar("test/test_accuracy", epoch_accuracy.numpy(), epoch)

        print(np.array(test_acc))
        print('目前准确率最大为' + str(tf.reduce_max(test_acc).numpy()))
        print(str(epoch_accuracy.numpy()), str(tf.reduce_max(test_acc).numpy()))
        # print(tr_acc.tolist())
        x = tr_acc - tf.reduce_max(Train_acc)
        # print(x.numpy())
        if x >= 0:
            Time = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())  # 本次循环开始时间，放到文件命名
            for i in file:  # 删除上一个模型文件，保存新的模型
                if os.path.exists(i):
                    os.remove(i)
                else:
                    print('no such file:%s' % i)
            print('*************************模型保存***************************************')
            discriminator.save_weights('Gan_model/model_time[{}]'.format(Time))
            file = ['D:/python/ssgan_tf2.0/Gan_model/model_time[{}].index'.format(Time),
                    'D:/python/ssgan_tf2.0/Gan_model/model_time[{}].data-00000-of-00002'.format(Time),
                    'D:/python/ssgan_tf2.0/Gan_model/model_time[{}].data-00001-of-00002'.format(Time)]
        test_acc.append(test_accuracy)
    del discriminator

    return train_hist, Train_acc


if __name__ == '__main__':
    Learning_rate = 0.0005
    Epochs = 100
    train_loss, train_acc = main(Learning_rate, Epochs)
    Draw(train_loss, Learning_rate, Epochs, show=True, save=True)
    Draw(train_acc, Learning_rate, Epochs, show=True, save=True, is_loss=False)

    # Path = 'D:/python/ssgan_tf2.0/accuracy.xlsx'
    # write_excel(Path, Learning_rate, 2, train_acc)
